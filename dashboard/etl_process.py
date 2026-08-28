# -*- coding: utf-8 -*-
"""
天宏平台车辆收款明细数据 ETL 与分析聚合引擎
"""
import os
import sys
import json
import datetime
import pandas as pd
import openpyxl

EXCEL_PATH = r"D:\shoukuan\2026年北斗平台服务车辆收款明细表5月.xlsx"
OUTPUT_DIR = r"D:\shoukuan\dashboard\data"
OUTPUT_JSON = os.path.join(OUTPUT_DIR, "dashboard_data.json")

def format_date(val):
    if val is None or val == "" or pd.isna(val):
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        try:
            if 20000 < val < 60000:
                base = datetime.datetime(1899, 12, 30)
                dt = base + datetime.timedelta(days=float(val))
                return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
    val_str = str(val).strip()
    if len(val_str) >= 10 and val_str[:10].count("-") == 2:
        return val_str[:10]
    return val_str

def parse_num(val):
    if val is None or pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    try:
        clean = str(val).replace(",", "").strip()
        if clean == "" or clean == "-":
            return 0.0
        return round(float(clean), 2)
    except Exception:
        return 0.0

def categorize_biz_type(val):
    if not val:
        return "未分类"
    s = str(val).strip()
    if "重型货车" in s or "重货" in s or "货车" in s or "挂车" in s:
        return "重型货车"
    if "网约" in s or "出租车" in s:
        return "网约出租车"
    if "客运" in s or "班车" in s or "公交" in s:
        return "客运班车"
    if "驾培" in s or "驾校" in s:
        return "驾培车辆"
    if "第三方" in s or "代理" in s or "易流" in s:
        return "第三方接入"
    if "工程" in s or "挖掘机" in s or "泥头车" in s or "搅拌" in s:
        return "工程机械"
    return "其他业务"

def run_etl(excel_path=EXCEL_PATH, output_json=OUTPUT_JSON):
    print(f"正在加载 Excel 文件: {excel_path} ...")
    os.makedirs(os.path.dirname(output_json), exist_ok=True)
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "天宏平台" not in wb.sheetnames:
        raise ValueError("工作簿中未找到「天宏平台」工作表！")
    
    sheet = wb["天宏平台"]
    print(f"「天宏平台」工作表行数: {sheet.max_row}, 列数: {sheet.max_column}")
    
    records = []
    base_date = datetime.date(2026, 5, 31)

    for r in range(5, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=j).value for j in range(1, 30)]
        if not any(v is not None for v in row_vals):
            continue
        
        plate = str(row_vals[0]).strip() if row_vals[0] is not None else ""
        org = str(row_vals[1]).strip() if row_vals[1] is not None else ""
        biz = str(row_vals[2]).strip() if row_vals[2] is not None else ""
        mgr = str(row_vals[3]).strip() if row_vals[3] is not None else "未指定"
        
        mgr = mgr.replace("（飞）", "").strip()
        if mgr in ["飞哥", "吴川"]:
            mgr = "吴锡飞"
        if "黄奕翔" in mgr:
            mgr = "黄奕翔"
        
        # 服务费
        s_due = format_date(row_vals[4])
        s_rec = parse_num(row_vals[5])
        s_paid = parse_num(row_vals[6])
        s_date = format_date(row_vals[7])
        s_unrec = parse_num(row_vals[8])
        
        # 第三方收款
        t_rec = parse_num(row_vals[9])
        t_paid = parse_num(row_vals[10])
        t_date = format_date(row_vals[11])
        t_unrec = parse_num(row_vals[12])
        t_due = format_date(row_vals[13])
        
        # 设备
        d_name = str(row_vals[14]).strip() if row_vals[14] is not None else ""
        d_inst = format_date(row_vals[15])
        d_rec = parse_num(row_vals[16])
        d_paid = parse_num(row_vals[17])
        d_date = format_date(row_vals[18])
        d_unrec = parse_num(row_vals[19])
        
        # 售后
        a_item = str(row_vals[20]).strip() if row_vals[20] is not None else ""
        a_rec = parse_num(row_vals[21])
        a_paid = parse_num(row_vals[22])
        a_date = format_date(row_vals[23])
        a_unrec = parse_num(row_vals[24])
        
        # 其他收款
        o_paid = parse_num(row_vals[25])
        o_date = format_date(row_vals[26])
        o_unrec = parse_num(row_vals[27])
        remark = str(row_vals[28]).strip() if row_vals[28] is not None else ""
        
        # 计算合计数
        total_rec = round(s_rec + t_rec + d_rec + a_rec, 2)
        total_paid = round(s_paid + t_paid + d_paid + a_paid + o_paid, 2)
        total_unrec = round(s_unrec + t_unrec + d_unrec + a_unrec + o_unrec, 2)
        
        # 收款状态判定
        if total_rec > 0 and total_unrec == 0 and total_paid >= total_rec:
            status = "已结清"
        elif total_paid > 0 and total_unrec > 0:
            status = "部分收款"
        elif total_paid == 0 and (total_unrec > 0 or total_rec > 0):
            status = "未交款"
        elif total_paid > 0 and total_rec == 0:
            status = "其他/溢收款"
        else:
            status = "无应收"
        
        # 到期状态判定
        effective_due = s_due if s_due else t_due
        expiry_status = "无服务期"
        days_to_expire = None
        if effective_due and len(effective_due) >= 10:
            try:
                due_d = datetime.datetime.strptime(effective_due[:10], "%Y-%m-%d").date()
                delta_days = (due_d - base_date).days
                days_to_expire = delta_days
                if delta_days < 0:
                    expiry_status = "已过期"
                elif delta_days <= 30:
                    expiry_status = "30天内到期"
                elif delta_days <= 90:
                    expiry_status = "90天内到期"
                else:
                    expiry_status = "正常服务中"
            except Exception:
                expiry_status = "无服务期"
        
        biz_cat = categorize_biz_type(biz)
        
        record = {
            "id": len(records) + 1,
            "plate_no": plate,
            "org_name": org,
            "biz_type": biz,
            "biz_category": biz_cat,
            "manager": mgr,
            "total_receivable": total_rec,
            "total_received": total_paid,
            "total_unreceived": total_unrec,
            "payment_status": status,
            "expiry_status": expiry_status,
            "days_to_expire": days_to_expire,
            "primary_due_date": effective_due,
            "service_due_date": s_due,
            "service_receivable": s_rec,
            "service_received": s_paid,
            "service_received_date": s_date,
            "service_unreceived": s_unrec,
            "third_receivable": t_rec,
            "third_received": t_paid,
            "third_received_date": t_date,
            "third_unreceived": t_unrec,
            "third_due_date": t_due,
            "device_name": d_name,
            "device_install_date": d_inst,
            "device_receivable": d_rec,
            "device_received": d_paid,
            "device_received_date": d_date,
            "device_unreceived": d_unrec,
            "aftersales_item": a_item,
            "aftersales_receivable": a_rec,
            "aftersales_received": a_paid,
            "aftersales_received_date": a_date,
            "aftersales_unreceived": a_unrec,
            "other_received": o_paid,
            "other_received_date": o_date,
            "other_unreceived": o_unrec,
            "remark": remark
        }
        records.append(record)

    df = pd.DataFrame(records)
    print(f"清洗完成，共提取有效记录: {len(df)} 条")

    # 1. 核心 KPI
    total_receivable = round(float(df["total_receivable"].sum()), 2)
    total_received = round(float(df["total_received"].sum()), 2)
    total_unreceived = round(float(df["total_unreceived"].sum()), 2)
    overall_rate = round(total_received / total_receivable * 100, 2) if total_receivable > 0 else 0.0
    
    vehicles_count = int(df[df["plate_no"] != ""]["plate_no"].nunique())
    orgs_count = int(df[df["org_name"] != ""]["org_name"].nunique())
    managers_count = int(df[df["manager"] != ""]["manager"].nunique())
    
    status_counts = df["payment_status"].value_counts().to_dict()
    expiry_counts = df["expiry_status"].value_counts().to_dict()
    
    kpis = {
        "total_records": len(df),
        "total_vehicles": vehicles_count,
        "total_orgs": orgs_count,
        "total_managers": managers_count,
        "total_receivable": total_receivable,
        "total_received": total_received,
        "total_unreceived": total_unreceived,
        "collection_rate": overall_rate,
        "cleared_count": int(status_counts.get("已结清", 0)),
        "partial_count": int(status_counts.get("部分收款", 0)),
        "unpaid_count": int(status_counts.get("未交款", 0)),
        "expired_count": int(expiry_counts.get("已过期", 0)),
        "expiring_30_count": int(expiry_counts.get("30天内到期", 0)),
        "normal_service_count": int(expiry_counts.get("正常服务中", 0))
    }

    # 2. 五大业务板块收支对比
    streams = [
        {
            "name": "服务费",
            "receivable": round(float(df["service_receivable"].sum()), 2),
            "received": round(float(df["service_received"].sum()), 2),
            "unreceived": round(float(df["service_unreceived"].sum()), 2),
            "rate": round(float(df["service_received"].sum()) / float(df["service_receivable"].sum()) * 100, 2) if df["service_receivable"].sum() > 0 else 0
        },
        {
            "name": "第三方收款",
            "receivable": round(float(df["third_receivable"].sum()), 2),
            "received": round(float(df["third_received"].sum()), 2),
            "unreceived": round(float(df["third_unreceived"].sum()), 2),
            "rate": round(float(df["third_received"].sum()) / float(df["third_receivable"].sum()) * 100, 2) if df["third_receivable"].sum() > 0 else 0
        },
        {
            "name": "设备费",
            "receivable": round(float(df["device_receivable"].sum()), 2),
            "received": round(float(df["device_received"].sum()), 2),
            "unreceived": round(float(df["device_unreceived"].sum()), 2),
            "rate": round(float(df["device_received"].sum()) / float(df["device_receivable"].sum()) * 100, 2) if df["device_receivable"].sum() > 0 else 0
        },
        {
            "name": "售后费",
            "receivable": round(float(df["aftersales_receivable"].sum()), 2),
            "received": round(float(df["aftersales_received"].sum()), 2),
            "unreceived": round(float(df["aftersales_unreceived"].sum()), 2),
            "rate": round(float(df["aftersales_received"].sum()) / float(df["aftersales_receivable"].sum()) * 100, 2) if df["aftersales_receivable"].sum() > 0 else 0
        },
        {
            "name": "其他收款",
            "receivable": 0.0,
            "received": round(float(df["other_received"].sum()), 2),
            "unreceived": round(float(df["other_unreceived"].sum()), 2),
            "rate": 100.0 if df["other_received"].sum() > 0 else 0
        }
    ]

    # 3. 业务负责人业绩统计
    mgr_group = df.groupby("manager").agg({
        "id": "count",
        "total_receivable": "sum",
        "total_received": "sum",
        "total_unreceived": "sum"
    }).reset_index()
    
    mgr_stats = []
    for _, row in mgr_group.iterrows():
        m_name = str(row["manager"]).strip()
        if not m_name or m_name == "nan":
            continue
        rec = round(float(row["total_receivable"]), 2)
        paid = round(float(row["total_received"]), 2)
        unrec = round(float(row["total_unreceived"]), 2)
        rate = round(paid / rec * 100, 2) if rec > 0 else (100.0 if paid > 0 else 0.0)
        mgr_stats.append({
            "manager": m_name,
            "vehicle_count": int(row["id"]),
            "receivable": rec,
            "received": paid,
            "unreceived": unrec,
            "rate": rate
        })
    mgr_stats.sort(key=lambda x: x["received"], reverse=True)

    # 4. 业务类型大类统计
    biz_group = df.groupby("biz_category").agg({
        "id": "count",
        "total_receivable": "sum",
        "total_received": "sum",
        "total_unreceived": "sum"
    }).reset_index()
    
    biz_stats = []
    for _, row in biz_group.iterrows():
        b_name = str(row["biz_category"]).strip()
        rec = round(float(row["total_receivable"]), 2)
        paid = round(float(row["total_received"]), 2)
        unrec = round(float(row["total_unreceived"]), 2)
        rate = round(paid / rec * 100, 2) if rec > 0 else 0
        biz_stats.append({
            "category": b_name,
            "count": int(row["id"]),
            "receivable": rec,
            "received": paid,
            "unreceived": unrec,
            "rate": rate
        })
    biz_stats.sort(key=lambda x: x["receivable"], reverse=True)

    # 5. Top 20 欠款大客户
    org_debt_group = df[df["total_unreceived"] > 0].groupby("org_name").agg({
        "id": "count",
        "total_receivable": "sum",
        "total_received": "sum",
        "total_unreceived": "sum"
    }).reset_index()
    
    org_debt_stats = []
    for _, row in org_debt_group.iterrows():
        o_name = str(row["org_name"]).strip()
        if not o_name or o_name == "nan":
            continue
        org_debt_stats.append({
            "org_name": o_name,
            "debt_vehicles": int(row["id"]),
            "receivable": round(float(row["total_receivable"]), 2),
            "received": round(float(row["total_received"]), 2),
            "unreceived": round(float(row["total_unreceived"]), 2)
        })
    org_debt_stats.sort(key=lambda x: x["unreceived"], reverse=True)
    top20_debtors = org_debt_stats[:20]

    # 6. 服务期到期时间分布
    year_map = {}
    for r in records:
        d_str = r["primary_due_date"]
        if d_str and len(d_str) >= 4:
            yr = d_str[:4]
            if yr.isdigit():
                year_map[yr] = year_map.get(yr, 0) + 1
            else:
                year_map["其他"] = year_map.get("其他", 0) + 1
        else:
            year_map["无日期"] = year_map.get("无日期", 0) + 1
    
    timeline_stats = [{"year": k, "count": v} for k, v in sorted(year_map.items())]

    # 7. 下拉选项
    all_managers = sorted(list(set(r["manager"] for r in records if r["manager"])))
    all_biz_types = sorted(list(set(r["biz_category"] for r in records if r["biz_category"])))
    
    dashboard_payload = {
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "kpis": kpis,
        "streams": streams,
        "managers": mgr_stats,
        "biz_categories": biz_stats,
        "top_debtors": top20_debtors,
        "timeline": timeline_stats,
        "options": {
            "managers": all_managers,
            "biz_categories": all_biz_types,
            "payment_statuses": ["全部", "已结清", "部分收款", "未交款", "其他/溢收款", "无应收"],
            "expiry_statuses": ["全部", "已过期", "30天内到期", "90天内到期", "正常服务中", "无服务期"]
        },
        "records": records
    }

    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(dashboard_payload, f, ensure_ascii=False)
    
    print(f"数据处理成功！已生成 JSON 数据集: {output_json} (大小: {os.path.getsize(output_json) / 1024 / 1024:.2f} MB)")
    return dashboard_payload

if __name__ == "__main__":
    run_etl()
